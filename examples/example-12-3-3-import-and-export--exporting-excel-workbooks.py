#!/usr/bin/python3
"""
-- Applied pandas
---- Imports and exports
------ Reading from and writing to Excel workbooks
-------- Exporting Excel workbooks
"""

import pandas as pd

baby_names = pd.read_csv("../datasets/new_york_city_baby_names.csv")

girls = baby_names[baby_names["Gender"] == "FEMALE"]
boys = baby_names[baby_names["Gender"] == "MALE"]

# Creating an ExcelWriter object also creates a file
writer = pd.ExcelWriter("../datasets/Baby_Names.xlsx")
print(writer)  #<pandas.io.excel._openpyxl.OpenpyxlWriter object at 0x000001A333EC09A0>

girls.to_excel(excel_writer=writer, sheet_name="Girls", index=False)
# The worksheet is not added yet to the file

boys.to_excel(
    writer,
    sheet_name="Boys",
    index=False,
    columns=["Child's First Name", "Count", "Rank"]
)
# The worksheet is not added yet to the file

# Now save in the file
writer.save()

# Use close() instead of save() in newer Pandas versions:
# writer.close()

# You can also use the 'with' context manager:
"""
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

with pd.ExcelWriter(to_file) as writer:
    df.to_excel(excel_writer=writer)
    workbook = writer.book
    worksheet = workbook.active
    first_column_letter = get_column_letter(1)
    worksheet.column_dimensions[first_column_letter].auto_size = True
    for row_num in range(2, worksheet.max_row + 1):
        worksheet[first_column_letter + str(row_num)].alignment = Alignment(horizontal="left")
    for col_num in range(2, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 20
"""
